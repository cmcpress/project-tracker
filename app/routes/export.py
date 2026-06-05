"""
routes/export.py — Export and import endpoints.
Baselines and critical path are also handled here.
Full PDF/Gantt export implementation in Phase 7.
"""
from __future__ import annotations


from flask import Blueprint, jsonify, request, abort, Response
from app.database import get_connection
from app.models import rows_to_list, row_to_dict
from datetime import datetime

bp = Blueprint("export", __name__, url_prefix="/api")


# ---------------------------------------------------------------------------
# Critical path
# ---------------------------------------------------------------------------

@bp.route("/projects/<int:project_id>/critical-path", methods=["GET"])
def get_critical_path(project_id: int):
    """
    Calculate and return the list of task IDs on the critical path.

    Algorithm: forward pass (ES/EF) + backward pass (LS/LF) over the DAG
    of FS dependencies. Tasks with float == 0 are on the critical path.
    Only tasks with both start_date and end_date are included.
    """
    conn = get_connection()
    try:
        if not conn.execute("SELECT id FROM projects WHERE id = ?", (project_id,)).fetchone():
            abort(404, f"Project {project_id} not found.")

        # Load tasks with dates
        task_rows = conn.execute(
            """SELECT id, start_date, end_date, duration_days
               FROM tasks
               WHERE project_id = ? AND start_date IS NOT NULL AND end_date IS NOT NULL""",
            (project_id,)
        ).fetchall()

        if not task_rows:
            return jsonify({"critical_path": []})

        tasks = {r["id"]: dict(r) for r in task_rows}
        task_ids = set(tasks.keys())

        # Compute duration in days for each task (use duration_days or derive from dates)
        from datetime import date
        for t in tasks.values():
            if t.get("duration_days"):
                t["dur"] = t["duration_days"]
            else:
                try:
                    s = date.fromisoformat(t["start_date"])
                    e = date.fromisoformat(t["end_date"])
                    t["dur"] = max(1, (e - s).days + 1)
                except ValueError:
                    t["dur"] = 1

        # Load FS dependencies between tasks in this project
        dep_rows = conn.execute(
            """SELECT d.predecessor_id, d.successor_id, d.lag_days
               FROM dependencies d
               JOIN tasks pt ON pt.id = d.predecessor_id
               JOIN tasks st ON st.id = d.successor_id
               WHERE pt.project_id = ? AND st.project_id = ? AND d.type = 'FS'""",
            (project_id, project_id)
        ).fetchall()

        # Adjacency: successors[task_id] = list of (successor_id, lag_days)
        # predecessors[task_id] = list of (predecessor_id, lag_days)
        successors = {tid: [] for tid in task_ids}
        predecessors = {tid: [] for tid in task_ids}
        for d in dep_rows:
            if d["predecessor_id"] in task_ids and d["successor_id"] in task_ids:
                successors[d["predecessor_id"]].append((d["successor_id"], d["lag_days"]))
                predecessors[d["successor_id"]].append((d["predecessor_id"], d["lag_days"]))

        # Forward pass — compute ES (Early Start) and EF (Early Finish)
        es = {tid: 0 for tid in task_ids}
        ef = {tid: tasks[tid]["dur"] for tid in task_ids}

        # Topological order (Kahn's algorithm)
        in_degree = {tid: len(predecessors[tid]) for tid in task_ids}
        queue = [tid for tid in task_ids if in_degree[tid] == 0]
        topo_order = []

        while queue:
            node = queue.pop(0)
            topo_order.append(node)
            for (succ, lag) in successors[node]:
                ef_node = ef[node] + lag
                if ef_node > es[succ]:
                    es[succ] = ef_node
                    ef[succ] = es[succ] + tasks[succ]["dur"]
                in_degree[succ] -= 1
                if in_degree[succ] == 0:
                    queue.append(succ)

        # If there are cycles, return empty (should not happen with valid data)
        if len(topo_order) != len(task_ids):
            return jsonify({"critical_path": [], "warning": "Dependency cycle detected"})

        # Backward pass — compute LF (Late Finish) and LS (Late Start)
        project_ef = max(ef.values())
        lf = {tid: project_ef for tid in task_ids}
        ls = {tid: project_ef - tasks[tid]["dur"] for tid in task_ids}

        for node in reversed(topo_order):
            for (succ, lag) in successors[node]:
                lf_via_succ = ls[succ] - lag
                if lf_via_succ < lf[node]:
                    lf[node] = lf_via_succ
                    ls[node] = lf[node] - tasks[node]["dur"]

        # Float = LS - ES; tasks with float == 0 are on the critical path
        critical = [tid for tid in task_ids if (ls[tid] - es[tid]) == 0]
        return jsonify({"critical_path": critical})
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Export / Import (Phase 7)
# ---------------------------------------------------------------------------

@bp.route("/export/project/<int:project_id>/pdf", methods=["POST"])
def export_pdf(project_id: int):
    """
    Generate a Gantt chart PDF using reportlab's canvas API.

    Renders: task name labels (with WBS indentation), month date header,
    task bars (coloured by status), today line, and a legend.
    Supports multi-page output for large task lists.
    """
    try:
        import io
        from datetime import date, timedelta
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A3, landscape
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas as pdf_canvas
    except ImportError:
        return jsonify({"error": "reportlab is not installed on this system."}), 501

    conn = get_connection()
    try:
        project_row = conn.execute(
            "SELECT * FROM projects WHERE id = ?", (project_id,)
        ).fetchone()
        if not project_row:
            abort(404, "Project not found")
        project = row_to_dict(project_row)

        tasks_rows = conn.execute(
            "SELECT * FROM tasks WHERE project_id = ? ORDER BY sort_order, id",
            (project_id,)
        ).fetchall()
        tasks = [row_to_dict(t) for t in tasks_rows]

        # ── Hierarchy helpers ────────────────────────────────────────────────
        parent_map = {t["id"]: t.get("parent_id") for t in tasks}

        def depth_of(task_id):
            d, pid = 0, parent_map.get(task_id)
            while pid is not None:
                d += 1
                pid = parent_map.get(pid)
            return min(d, 5)

        # ── Date range ───────────────────────────────────────────────────────
        today = date.today()
        valid_starts, valid_ends = [], []
        for t in tasks:
            for field, lst in (("start_date", valid_starts), ("end_date", valid_ends)):
                val = t.get(field)
                if val:
                    try:
                        lst.append(date.fromisoformat(val))
                    except ValueError:
                        pass

        if not valid_starts:
            valid_starts = [today]
        if not valid_ends:
            valid_ends = [today + timedelta(days=30)]

        chart_start = min(valid_starts) - timedelta(days=3)
        chart_end   = max(valid_ends)   + timedelta(days=3)
        total_days  = (chart_end - chart_start).days + 1

        # First of month helper
        def next_month_start(d):
            if d.month == 12:
                return d.replace(year=d.year + 1, month=1, day=1)
            return d.replace(month=d.month + 1, day=1)

        # ── Page layout constants ────────────────────────────────────────────
        PAGE_W, PAGE_H = landscape(A3)   # 841.9 × 595.3 pt  (≈ 297mm × 210mm)
        MX   = 10 * mm    # left/right margin
        MT   = 10 * mm    # top margin
        MB   = 12 * mm    # bottom margin (room for legend)
        LW   = 78 * mm    # label column width
        TH   = 16 * mm    # title block height
        HH   =  9 * mm    # date header height
        RH   =  7 * mm    # row height

        chart_avail  = PAGE_W - 2 * MX - LW
        day_w        = min(chart_avail / total_days, 5 * mm)
        chart_w      = day_w * total_days   # actual drawn width (≤ chart_avail)

        rows_avail   = PAGE_H - MT - TH - HH - MB
        n_per_page   = max(1, int(rows_avail / RH))

        # ── Status colours ───────────────────────────────────────────────────
        STATUS_COLOR = {
            "not-started": colors.HexColor("#9ca3af"),
            "planning":    colors.HexColor("#3b82f6"),
            "in-progress": colors.HexColor("#f97316"),
            "blocked":     colors.HexColor("#ef4444"),
            "complete":    colors.HexColor("#22c55e"),
        }
        STATUS_LABEL = {
            "not-started": "Not started",
            "planning":    "Planning",
            "in-progress": "In progress",
            "blocked":     "Blocked",
            "complete":    "Complete",
        }

        def date_x(d):
            """Convert a date to x-coordinate in points."""
            return MX + LW + (d - chart_start).days * day_w

        # ── Draw one page ────────────────────────────────────────────────────
        def draw_page(cv, page_rows, page_num, total_pages):
            # ── Title block ──────────────────────────────────────────────────
            title_y = PAGE_H - MT
            cv.setFont("Helvetica-Bold", 12)
            cv.setFillColor(colors.HexColor("#1f2937"))
            cv.drawString(MX, title_y - 10, project["name"])

            cv.setFont("Helvetica", 7)
            cv.setFillColor(colors.HexColor("#6b7280"))
            meta = f"Generated {today.strftime('%d %b %Y')}"
            if total_pages > 1:
                meta += f"   ·   Page {page_num} of {total_pages}"
            cv.drawString(MX, title_y - 19, meta)

            # ── Date header ──────────────────────────────────────────────────
            hdr_top = PAGE_H - MT - TH
            hdr_bot = hdr_top - HH

            # Month band backgrounds + labels
            cur = chart_start.replace(day=1)
            while cur <= chart_end:
                nxt = next_month_start(cur)
                x1  = max(date_x(cur),   date_x(chart_start))
                x2  = min(date_x(nxt),   date_x(chart_end) + day_w)
                if x2 > x1:
                    bg = "#f0f2f5" if cur.month % 2 == 0 else "#ffffff"
                    cv.setFillColor(colors.HexColor(bg))
                    cv.rect(x1, hdr_bot, x2 - x1, HH, fill=1, stroke=0)

                    cv.setFont("Helvetica-Bold", 6.5)
                    cv.setFillColor(colors.HexColor("#374151"))
                    cv.drawString(x1 + 1.5, hdr_bot + 2.5 * mm,
                                  cur.strftime("%b %Y"))

                    # Month separator line
                    if cur > chart_start.replace(day=1):
                        cv.setStrokeColor(colors.HexColor("#d1d5db"))
                        cv.setLineWidth(0.3)
                        cv.line(x1, hdr_top, x1, hdr_bot)
                cur = nxt

            # Header border
            cv.setStrokeColor(colors.HexColor("#d1d5db"))
            cv.setLineWidth(0.4)
            cv.rect(MX + LW, hdr_bot, chart_w, HH, fill=0, stroke=1)
            # Label column header
            cv.rect(MX, hdr_bot, LW, HH, fill=0, stroke=1)
            cv.setFont("Helvetica-Bold", 6.5)
            cv.setFillColor(colors.HexColor("#374151"))
            cv.drawString(MX + 2, hdr_bot + 2.5 * mm, "Task")

            # Today tick in header
            if chart_start <= today <= chart_end:
                tx = date_x(today)
                cv.setStrokeColor(colors.HexColor("#e53935"))
                cv.setLineWidth(1)
                cv.line(tx, hdr_top, tx, hdr_bot)

            # ── Task rows ────────────────────────────────────────────────────
            bottom_y = hdr_bot  # will be updated as rows are drawn

            for ri, task in enumerate(page_rows):
                ry = hdr_bot - (ri + 1) * RH
                bottom_y = ry
                is_group = task.get("type") == "group"
                depth    = depth_of(task["id"])

                # Alternating row background
                if ri % 2 == 1:
                    cv.setFillColor(colors.HexColor("#f8f9fa"))
                    cv.rect(MX, ry, LW + chart_w, RH, fill=1, stroke=0)

                # Row bottom border
                cv.setStrokeColor(colors.HexColor("#e4e7ec"))
                cv.setLineWidth(0.2)
                cv.line(MX, ry, MX + LW + chart_w, ry)

                # Month grid lines (vertical, chart area only)
                cur = chart_start.replace(day=1)
                while cur <= chart_end:
                    if cur > chart_start.replace(day=1):
                        gx = date_x(cur)
                        if MX + LW < gx < MX + LW + chart_w:
                            cv.setStrokeColor(colors.HexColor("#e4e7ec"))
                            cv.setLineWidth(0.2)
                            cv.line(gx, ry, gx, ry + RH)
                    cur = next_month_start(cur)

                # Today row line
                if chart_start <= today <= chart_end:
                    tx = date_x(today)
                    cv.setStrokeColor(colors.HexColor("#fca5a5"))
                    cv.setLineWidth(0.4)
                    cv.line(tx, ry, tx, ry + RH)

                # Task label
                indent = MX + depth * 3 * mm + 2 * mm
                font   = "Helvetica-Bold" if is_group else "Helvetica"
                fsize  = 6.5
                cv.setFont(font, fsize)
                cv.setFillColor(colors.HexColor("#1f2937"))
                name     = task.get("name") or ""
                max_w    = LW - depth * 3 * mm - 4 * mm
                # Truncate to fit
                while name and cv.stringWidth(name, font, fsize) > max_w:
                    name = name[:-1]
                if len(name) < len(task.get("name") or ""):
                    name = name[:-1] + "…"
                cv.drawString(indent, ry + 2.2 * mm, name)

                # Task bar
                t_start = task.get("start_date")
                t_end   = task.get("end_date")
                if t_start and t_end:
                    try:
                        ts = date.fromisoformat(t_start)
                        te = date.fromisoformat(t_end)
                        # Clip to chart range
                        bx1 = date_x(max(ts, chart_start))
                        bx2 = date_x(min(te, chart_end)) + day_w
                        bw  = max(bx2 - bx1, 1.5)

                        status    = task.get("status") or "not-started"
                        bar_color = STATUS_COLOR.get(status, STATUS_COLOR["not-started"])
                        pad_v     = 1.2 * mm
                        bar_y     = ry + pad_v
                        bar_h     = RH - 2 * pad_v

                        if is_group:
                            # Thinner dark bar for group tasks
                            cv.setFillColor(colors.HexColor("#475569"))
                            cv.rect(bx1, bar_y + bar_h * 0.3, bw, bar_h * 0.4,
                                    fill=1, stroke=0)
                        else:
                            cv.setFillColor(bar_color)
                            cv.roundRect(bx1, bar_y, bw, bar_h,
                                         radius=1.5, fill=1, stroke=0)

                            # Progress marker (vertical white line)
                            progress = float(task.get("progress") or 0)
                            if 0 < progress < 1 and status != "complete":
                                prog_x = bx1 + bw * progress
                                cv.setStrokeColor(colors.white)
                                cv.setLineWidth(1.2)
                                cv.line(prog_x, bar_y + 1, prog_x, bar_y + bar_h - 1)

                    except (ValueError, TypeError):
                        pass

            # ── Frame / borders ──────────────────────────────────────────────
            frame_bot = bottom_y if page_rows else hdr_bot
            cv.setStrokeColor(colors.HexColor("#d1d5db"))
            cv.setLineWidth(0.5)
            cv.line(MX,           hdr_top, MX,                frame_bot)  # left
            cv.line(MX + LW,      hdr_top, MX + LW,           frame_bot)  # separator
            cv.line(MX + LW + chart_w, hdr_top, MX + LW + chart_w, frame_bot)  # right
            cv.line(MX,           frame_bot, MX + LW + chart_w, frame_bot)  # bottom

            # Today full dashed line
            if chart_start <= today <= chart_end:
                tx = date_x(today)
                cv.setStrokeColor(colors.HexColor("#e53935"))
                cv.setLineWidth(0.7)
                cv.setDash(3, 3)
                cv.line(tx, hdr_top, tx, frame_bot)
                cv.setDash()

            # ── Legend (bottom of page) ───────────────────────────────────────
            if page_num == total_pages:
                lx = MX
                ly = MB * 0.4
                cv.setFont("Helvetica", 6)
                cv.setFillColor(colors.HexColor("#6b7280"))
                cv.drawString(lx, ly + 3, "Status: ")
                lx += cv.stringWidth("Status: ", "Helvetica", 6) + 1
                for key, label in STATUS_LABEL.items():
                    swatch_col = STATUS_COLOR[key]
                    cv.setFillColor(swatch_col)
                    cv.rect(lx, ly, 7, 7, fill=1, stroke=0)
                    cv.setFillColor(colors.HexColor("#374151"))
                    cv.drawString(lx + 9, ly + 1, label)
                    lx += 9 + cv.stringWidth(label, "Helvetica", 6) + 8

                # Today marker in legend
                cv.setStrokeColor(colors.HexColor("#e53935"))
                cv.setLineWidth(1)
                cv.setDash(3, 2)
                cv.line(lx + 2, ly, lx + 2, ly + 7)
                cv.setDash()
                cv.setFillColor(colors.HexColor("#374151"))
                cv.setFont("Helvetica", 6)
                cv.drawString(lx + 6, ly + 1, "Today")

        # ── Render all pages ─────────────────────────────────────────────────
        buf = io.BytesIO()
        cv  = pdf_canvas.Canvas(buf, pagesize=landscape(A3))

        total_pages = max(1, -(-len(tasks) // n_per_page))  # ceiling division

        for page_num in range(1, total_pages + 1):
            start_i    = (page_num - 1) * n_per_page
            page_rows  = tasks[start_i: start_i + n_per_page]
            draw_page(cv, page_rows, page_num, total_pages)
            if page_num < total_pages:
                cv.showPage()

        cv.save()
        pdf_bytes = buf.getvalue()

        safe_name = "".join(
            ch if ch.isalnum() or ch in (" ", "-", "_") else "_"
            for ch in project["name"]
        ).strip() or "project"

        return Response(
            pdf_bytes,
            mimetype="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="{safe_name}-gantt.pdf"',
                "Content-Length": str(len(pdf_bytes)),
            }
        )
    finally:
        conn.close()


@bp.route("/export/project/<int:project_id>/gantt", methods=["GET"])
def export_gantt(project_id: int):
    """Export the Gantt chart as PNG. (Future enhancement.)"""
    return jsonify({"error": "Gantt PNG export not yet implemented"}), 501


# ---------------------------------------------------------------------------
# Per-project JSON export
# ---------------------------------------------------------------------------

@bp.route("/export/project/<int:project_id>/json", methods=["GET"])
def export_project_json(project_id: int):
    """
    Export a single project and all its data as a JSON backup.
    Same structure as /export/data but limited to one project.
    """
    conn = get_connection()
    try:
        p_row = conn.execute(
            "SELECT * FROM projects WHERE id = ?", (project_id,)
        ).fetchone()
        if p_row is None:
            abort(404, "Project not found.")
        p = row_to_dict(p_row)

        tasks_rows = conn.execute(
            "SELECT * FROM tasks WHERE project_id = ? ORDER BY sort_order, id",
            (project_id,)
        ).fetchall()
        tasks = []
        for t in tasks_rows:
            task = row_to_dict(t)
            task["is_firm_date"] = bool(task.get("is_firm_date"))
            task["assignee_ids"] = [
                r["person_id"] for r in conn.execute(
                    "SELECT person_id FROM task_people WHERE task_id = ?", (t["id"],)
                ).fetchall()
            ]
            task["items"] = rows_to_list(conn.execute(
                "SELECT * FROM task_items WHERE task_id = ? ORDER BY sort_order",
                (t["id"],)
            ).fetchall())
            tasks.append(task)

        deps = rows_to_list(conn.execute(
            """SELECT d.predecessor_id, d.successor_id, d.type, d.lag_days
               FROM dependencies d
               JOIN tasks t ON t.id = d.predecessor_id
               WHERE t.project_id = ?""", (project_id,)
        ).fetchall())

        phases = rows_to_list(conn.execute(
            "SELECT * FROM phases WHERE project_id = ? ORDER BY sort_order, id",
            (project_id,)
        ).fetchall())

        links = rows_to_list(conn.execute(
            "SELECT * FROM project_links WHERE project_id = ? ORDER BY sort_order, id",
            (project_id,)
        ).fetchall())

        # Include people assigned to tasks in this project
        people_ids = set()
        for t in tasks:
            people_ids.update(t.get("assignee_ids", []))
        people = []
        for pid in people_ids:
            person_row = conn.execute("SELECT * FROM people WHERE id = ?", (pid,)).fetchone()
            if person_row:
                people.append(row_to_dict(person_row))

        p["tasks"]        = tasks
        p["dependencies"] = deps
        p["phases"]       = phases
        p["links"]        = links

        payload = {
            "version":     1,
            "exported_at": datetime.utcnow().isoformat(),
            "people":      people,
            "projects":    [p],
        }

        import json
        ts = datetime.utcnow().strftime("%Y%m%d")
        safe_name = "".join(c for c in p["name"] if c.isalnum() or c in " -_")[:40].strip()
        filename = f"{safe_name}-{ts}.json"

        return Response(
            json.dumps(payload, indent=2),
            mimetype="application/json",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Per-project Excel export
# ---------------------------------------------------------------------------

@bp.route("/export/project/<int:project_id>/excel", methods=["GET"])
def export_project_excel(project_id: int):
    """Export a single project as an Excel workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import io
    except ImportError:
        return jsonify({"error": "openpyxl is not installed."}), 501

    conn = get_connection()
    try:
        p_row = conn.execute(
            "SELECT * FROM projects WHERE id = ?", (project_id,)
        ).fetchone()
        if p_row is None:
            abort(404, "Project not found.")
        p = row_to_dict(p_row)

        COLS = [
            ("WBS",            "wbs_number",        10),
            ("Name",           "name",              35),
            ("Type",           "type",              12),
            ("Status",         "status",            14),
            ("RAG",            "rag",                8),
            ("Progress (%)",   "progress",          12),
            ("Start Date",     "start_date",        13),
            ("End Date",       "end_date",          13),
            ("Act. Start",     "actual_start_date", 13),
            ("Act. End",       "actual_end_date",   13),
            ("Budget",         "budget",            12),
            ("Est. Hours",     "estimated_hours",   11),
            ("Logged Hours",   "logged_hours",      11),
            ("Notes",          "notes",             40),
        ]

        wb   = openpyxl.Workbook()
        ws   = wb.active
        ws.title = p["name"][:31].replace("/", "-")

        hdr_font  = Font(bold=True, color="FFFFFF")
        hdr_fill  = PatternFill("solid", fgColor="2563EB")
        ctr_align = Alignment(horizontal="center", vertical="center")

        # Project info block
        ws.append([f"Project: {p['name']}"])
        ws["A1"].font = Font(bold=True, size=13)
        ws.append([f"Category: {p.get('category','—')}   Status: {p.get('status','—')}"])
        if p.get("description"):
            ws.append([p["description"]])
        ws.append([f"Exported: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC"])
        ws.append([])

        hdr_row = ws.max_row + 1
        for ci, (label, _, _) in enumerate(COLS, 1):
            cell = ws.cell(row=hdr_row, column=ci, value=label)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = ctr_align

        tasks = conn.execute(
            "SELECT * FROM tasks WHERE project_id = ? ORDER BY sort_order, id",
            (project_id,)
        ).fetchall()

        for t in tasks:
            row = []
            for _, field, _ in COLS:
                val = t[field] if t[field] is not None else ""
                if field == "progress" and val != "":
                    val = round(float(val) * 100, 1)
                row.append(val)
            ws.append(row)

        for ci, (_, _, width) in enumerate(COLS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        ts = datetime.utcnow().strftime("%Y%m%d")
        safe_name = "".join(c for c in p["name"] if c.isalnum() or c in " -_")[:40].strip()

        return Response(
            buf.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}-{ts}.xlsx"'},
        )
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Per-project printable task report (PDF)
# ---------------------------------------------------------------------------

@bp.route("/export/project/<int:project_id>/report", methods=["GET"])
def export_project_report(project_id: int):
    """
    Generate a printable task report PDF for a single project.
    Portrait A4, formatted task list with summary block.
    Not a Gantt chart — a readable document for sharing.
    """
    try:
        import io
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas as pdf_canvas
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_LEFT, TA_CENTER
    except ImportError:
        return jsonify({"error": "reportlab is not installed."}), 501

    conn = get_connection()
    try:
        p_row = conn.execute(
            "SELECT * FROM projects WHERE id = ?", (project_id,)
        ).fetchone()
        if p_row is None:
            abort(404, "Project not found.")
        p = row_to_dict(p_row)

        tasks_rows = conn.execute(
            "SELECT * FROM tasks WHERE project_id = ? ORDER BY sort_order, id",
            (project_id,)
        ).fetchall()
        tasks = [row_to_dict(t) for t in tasks_rows]

        # Build assignee name lookup
        all_people = {
            r["id"]: r["name"] for r in
            conn.execute("SELECT id, name FROM people").fetchall()
        }
        for t in tasks:
            ids = [
                r["person_id"] for r in conn.execute(
                    "SELECT person_id FROM task_people WHERE task_id = ?", (t["id"],)
                ).fetchall()
            ]
            t["_assignees"] = ", ".join(all_people.get(i, "") for i in ids)

        # ── Stats ────────────────────────────────────────────────────────────
        total       = len(tasks)
        complete    = sum(1 for t in tasks if t.get("status") == "complete")
        overdue     = sum(1 for t in tasks
                         if t.get("end_date") and t.get("status") != "complete"
                         and t["end_date"] < datetime.utcnow().date().isoformat())
        total_budget = sum(float(t.get("budget") or 0) for t in tasks)
        total_est    = sum(float(t.get("estimated_hours") or 0) for t in tasks)
        total_logged = sum(float(t.get("logged_hours") or 0) for t in tasks)

        sym_row = conn.execute(
            "SELECT value FROM settings WHERE key = 'currency_symbol'"
        ).fetchone()
        sym = sym_row["value"] if sym_row else "£"

        # ── Build PDF ────────────────────────────────────────────────────────
        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=A4,
            leftMargin=15*mm, rightMargin=15*mm,
            topMargin=15*mm,  bottomMargin=15*mm,
        )

        styles = getSampleStyleSheet()
        h1  = ParagraphStyle("H1",  parent=styles["Heading1"],  fontSize=18, spaceAfter=2*mm)
        h2  = ParagraphStyle("H2",  parent=styles["Heading2"],  fontSize=11, spaceAfter=1*mm)
        sm  = ParagraphStyle("SM",  parent=styles["Normal"],    fontSize=8,  leading=11)
        smg = ParagraphStyle("SMG", parent=styles["Normal"],    fontSize=8,  leading=11,
                             textColor=colors.HexColor("#6b7280"))

        story = []

        # Title block
        story.append(Paragraph(p["name"], h1))
        meta_parts = [f"Category: {p.get('category','—')}", f"Status: {p.get('status','—').replace('-',' ').title()}"]
        story.append(Paragraph("  ·  ".join(meta_parts), smg))
        if p.get("description"):
            story.append(Spacer(1, 2*mm))
            story.append(Paragraph(p["description"], sm))
        story.append(Spacer(1, 3*mm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#e5e7eb")))
        story.append(Spacer(1, 3*mm))

        # Summary table
        rag_map = {"red": "🔴 Red", "amber": "🟡 Amber", "green": "🟢 Green"}
        summary_data = [
            ["Tasks", str(total),
             "Complete", f"{complete} ({round(complete/total*100) if total else 0}%)"],
            ["Overdue", str(overdue),
             "Budget", f"{sym}{total_budget:,.2f}" if total_budget else "—"],
            ["Est. hours", f"{total_est:g}h" if total_est else "—",
             "Logged hours", f"{total_logged:g}h" if total_logged else "—"],
        ]
        sum_tbl = Table(summary_data, colWidths=[30*mm, 25*mm, 30*mm, 35*mm])
        sum_tbl.setStyle(TableStyle([
            ("FONTSIZE",    (0,0), (-1,-1), 8),
            ("FONTNAME",    (0,0), (0,-1), "Helvetica-Bold"),
            ("FONTNAME",    (2,0), (2,-1), "Helvetica-Bold"),
            ("TEXTCOLOR",   (1,0), (1,-1), colors.HexColor("#111827")),
            ("TEXTCOLOR",   (3,0), (3,-1), colors.HexColor("#111827")),
            ("ROWBACKGROUNDS", (0,0), (-1,-1),
             [colors.HexColor("#f9fafb"), colors.white]),
            ("GRID",        (0,0), (-1,-1), 0.3, colors.HexColor("#e5e7eb")),
            ("TOPPADDING",  (0,0), (-1,-1), 3),
            ("BOTTOMPADDING",(0,0),(-1,-1), 3),
            ("LEFTPADDING", (0,0), (-1,-1), 4),
        ]))
        story.append(sum_tbl)
        story.append(Spacer(1, 5*mm))

        # Task table
        story.append(Paragraph("Tasks", h2))

        RAG_COLORS = {
            "red":   colors.HexColor("#ef4444"),
            "amber": colors.HexColor("#f59e0b"),
            "green": colors.HexColor("#22c55e"),
        }
        STATUS_COLORS = {
            "complete":    colors.HexColor("#22c55e"),
            "in-progress": colors.HexColor("#3b82f6"),
            "blocked":     colors.HexColor("#ef4444"),
            "pending":     colors.HexColor("#7c3aed"),
        }

        col_w = [12*mm, 55*mm, 16*mm, 20*mm, 10*mm, 18*mm, 18*mm, 18*mm]
        tbl_header = [["WBS", "Task Name", "Type", "Status", "RAG",
                        "Start", "End", "Assignees"]]
        tbl_data   = list(tbl_header)

        for t in tasks:
            indent = ""
            if t.get("parent_id"):
                indent = "  "
            name_str = indent + (t.get("name") or "")
            rag_str  = {"red": "■", "amber": "■", "green": "■"}.get(t.get("rag") or "", "")
            tbl_data.append([
                t.get("wbs_number") or "",
                Paragraph(name_str, sm),
                (t.get("type") or "task").capitalize(),
                (t.get("status") or "").replace("-", " ").title(),
                rag_str,
                t.get("start_date") or "",
                t.get("end_date") or "",
                Paragraph(t.get("_assignees") or "", sm),
            ])

        task_tbl = Table(tbl_data, colWidths=col_w, repeatRows=1)
        ts = [
            ("FONTSIZE",     (0,0), (-1,-1), 7.5),
            ("FONTNAME",     (0,0), (-1,0),  "Helvetica-Bold"),
            ("BACKGROUND",   (0,0), (-1,0),  colors.HexColor("#2563eb")),
            ("TEXTCOLOR",    (0,0), (-1,0),  colors.white),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),
             [colors.white, colors.HexColor("#f9fafb")]),
            ("GRID",         (0,0), (-1,-1), 0.3, colors.HexColor("#e5e7eb")),
            ("TOPPADDING",   (0,0), (-1,-1), 2),
            ("BOTTOMPADDING",(0,0), (-1,-1), 2),
            ("LEFTPADDING",  (0,0), (-1,-1), 3),
            ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ]

        # Colour RAG cells
        for row_i, t in enumerate(tasks, start=1):
            rag = t.get("rag")
            if rag in RAG_COLORS:
                ts.append(("TEXTCOLOR", (4, row_i), (4, row_i), RAG_COLORS[rag]))

        task_tbl.setStyle(TableStyle(ts))
        story.append(task_tbl)

        # Footer note
        story.append(Spacer(1, 5*mm))
        story.append(Paragraph(
            f"Generated {datetime.utcnow().strftime('%d %b %Y %H:%M')} UTC",
            smg
        ))

        doc.build(story)
        buf.seek(0)

        safe_name = "".join(c for c in p["name"] if c.isalnum() or c in " -_")[:40].strip()
        date_str  = datetime.utcnow().strftime("%Y%m%d")

        return Response(
            buf.read(),
            mimetype="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}-report-{date_str}.pdf"'},
        )
    finally:
        conn.close()


@bp.route("/export/data", methods=["GET"])
def export_data():
    """Export all data as a JSON backup."""
    conn = get_connection()
    try:
        people = rows_to_list(conn.execute("SELECT * FROM people ORDER BY id").fetchall())
        projects_rows = conn.execute("SELECT * FROM projects ORDER BY sort_order, id").fetchall()

        projects = []
        for p in projects_rows:
            proj = row_to_dict(p)
            tasks_rows = conn.execute(
                "SELECT * FROM tasks WHERE project_id = ? ORDER BY sort_order, id", (p["id"],)
            ).fetchall()
            tasks = []
            for t in tasks_rows:
                task = row_to_dict(t)
                task["is_firm_date"] = bool(task.get("is_firm_date"))
                assignee_ids = [
                    r["person_id"] for r in
                    conn.execute("SELECT person_id FROM task_people WHERE task_id = ?", (t["id"],)).fetchall()
                ]
                task["assignee_ids"] = assignee_ids
                items = rows_to_list(conn.execute(
                    "SELECT * FROM task_items WHERE task_id = ? ORDER BY sort_order", (t["id"],)
                ).fetchall())
                task["items"] = items
                tasks.append(task)

            deps = rows_to_list(conn.execute(
                """SELECT d.predecessor_id, d.successor_id, d.type, d.lag_days
                   FROM dependencies d
                   JOIN tasks t ON t.id = d.predecessor_id
                   WHERE t.project_id = ?""",
                (p["id"],)
            ).fetchall())

            phases = rows_to_list(conn.execute(
                "SELECT * FROM phases WHERE project_id = ? ORDER BY sort_order, id",
                (p["id"],)
            ).fetchall())

            proj["tasks"]        = tasks
            proj["dependencies"] = deps
            proj["phases"]       = phases
            projects.append(proj)

        # Attach unavailability to each person
        for p in people:
            p["unavailability"] = rows_to_list(conn.execute(
                "SELECT * FROM unavailability WHERE person_id = ? ORDER BY start_date, id",
                (p["id"],)
            ).fetchall())

        return jsonify({
            "version": 1,
            "exported_at": datetime.utcnow().isoformat(),
            "people": people,
            "projects": projects,
        })
    finally:
        conn.close()


@bp.route("/export/data/excel", methods=["GET"])
def export_excel():
    """Export all projects + tasks as an Excel workbook (one sheet per project)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import io
    except ImportError:
        return jsonify({"error": "openpyxl is not installed on this system."}), 501

    conn = get_connection()
    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default blank sheet

        header_font  = Font(bold=True, color="FFFFFF")
        header_fill  = PatternFill("solid", fgColor="2563EB")
        center_align = Alignment(horizontal="center", vertical="center")

        COLUMNS = [
            ("Name",           "name",              35),
            ("Type",           "type",              12),
            ("Status",         "status",            14),
            ("Start Date",     "start_date",        13),
            ("End Date",       "end_date",          13),
            ("Duration (days)","duration_days",     14),
            ("Actual Start",   "actual_start_date", 13),
            ("Actual End",     "actual_end_date",   13),
            ("Notes",          "notes",             40),
        ]

        projects_rows = conn.execute(
            "SELECT * FROM projects ORDER BY sort_order, id"
        ).fetchall()

        for p in projects_rows:
            # Sheet name: Excel limits to 31 chars, no special chars
            sheet_name = p["name"][:31].replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "").replace("[", "").replace("]", "").replace(":", "").strip()
            ws = wb.create_sheet(title=sheet_name or f"Project {p['id']}")

            # Project metadata rows
            ws.append([f"Project: {p['name']}"])
            ws["A1"].font = Font(bold=True, size=13)
            if p["description"]:
                ws.append([p["description"]])
            ws.append([f"Category: {p['category'] or '—'}"])
            ws.append([])  # blank spacer

            # Header row
            header_row = ws.max_row + 1
            for col_idx, (label, _, _) in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=header_row, column=col_idx, value=label)
                cell.font       = header_font
                cell.fill       = header_fill
                cell.alignment  = center_align

            # Task rows
            tasks = conn.execute(
                "SELECT * FROM tasks WHERE project_id = ? ORDER BY sort_order, id",
                (p["id"],)
            ).fetchall()
            for t in tasks:
                row_data = [t[field] if t[field] is not None else "" for _, field, _ in COLUMNS]
                ws.append(row_data)

            # Column widths
            for col_idx, (_, _, width) in enumerate(COLUMNS, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            # Freeze header row
            ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        ts = datetime.utcnow().strftime("%Y%m%d")
        return Response(
            buf.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="project-tracker-{ts}.xlsx"'},
        )
    finally:
        conn.close()


@bp.route("/import/data", methods=["POST"])
def import_data():
    """
    Import a JSON backup produced by /export/data.

    Strategy:
    - People   : match by name (case-insensitive); reuse existing or create new.
    - Projects : always create as new (no merge with existing data).
    - Tasks    : create new; maintain old_id → new_id mapping.
    - Dependencies : remap predecessor/successor using the task ID mapping.
    - task_people  : remap task_id and person_id.
    - task_items   : remap task_id.

    Returns: {"ok": true, "projects_imported": N, "tasks_imported": N}
    """
    import logging as _log
    logger = _log.getLogger(__name__)

    data = request.get_json(silent=True)
    if not data or not isinstance(data, dict):
        abort(400, "Request body must be a JSON object.")
    if data.get("version") != 1:
        abort(400, "Invalid backup format — expected {\"version\": 1}.")

    people_data   = data.get("people", [])
    projects_data = data.get("projects", [])

    conn = get_connection()
    try:
        # ── 1. People ─────────────────────────────────────────────────────────
        # Map old_person_id → new_person_id
        people_map: dict[int, int] = {}

        for person in people_data:
            old_id = person.get("id")
            name   = (person.get("name") or "").strip()
            if not name:
                continue
            existing = conn.execute(
                "SELECT id FROM people WHERE LOWER(name) = LOWER(?)", (name,)
            ).fetchone()
            if existing:
                people_map[old_id] = existing["id"]
            else:
                cur = conn.execute(
                    "INSERT INTO people (name, role, email, colour) VALUES (?,?,?,?)",
                    (name,
                     person.get("role"),
                     person.get("email"),
                     person.get("colour") or "#8892a4")
                )
                people_map[old_id] = cur.lastrowid

        # ── 1b. Unavailability (per-person, restored using the people_map) ──────
        for person in people_data:
            old_id   = person.get("id")
            new_id   = people_map.get(old_id)
            if not new_id:
                continue
            for u in (person.get("unavailability") or []):
                sd = (u.get("start_date") or "").strip()
                ed = (u.get("end_date")   or "").strip()
                if not sd or not ed:
                    continue
                conn.execute(
                    """INSERT INTO unavailability (person_id, start_date, end_date, label)
                       VALUES (?, ?, ?, ?)""",
                    (new_id, sd, ed, (u.get("label") or "Unavailable"))
                )

        # ── 2. Projects + tasks + dependencies ────────────────────────────────
        projects_imported = 0
        tasks_imported    = 0

        for proj in projects_data:
            proj_name = (proj.get("name") or "").strip()
            if not proj_name:
                continue

            # Determine max sort_order so imported projects go to the end
            max_order_row = conn.execute(
                "SELECT COALESCE(MAX(sort_order),0) FROM projects"
            ).fetchone()
            next_order = (max_order_row[0] or 0) + 1

            cur = conn.execute(
                """INSERT INTO projects
                   (name, category, status, description, colour, sort_order)
                   VALUES (?,?,?,?,?,?)""",
                (proj_name,
                 proj.get("category") or "General",
                 proj.get("status")   or "not-started",
                 proj.get("description"),
                 proj.get("colour")   or "#4a90e2",
                 next_order)
            )
            new_proj_id = cur.lastrowid
            projects_imported += 1

            # Task ID mapping for this project: old_task_id → new_task_id
            task_map: dict[int, int] = {}

            # Tasks — first pass: create all tasks without parent_id so tree
            # order doesn't matter. We remap parent_id in a second pass.
            for task in proj.get("tasks", []):
                old_task_id = task.get("id")
                task_name   = (task.get("name") or "").strip()
                if not task_name:
                    continue

                cur = conn.execute(
                    """INSERT INTO tasks (
                           project_id, name, type, status,
                           start_date, end_date, duration_days, is_firm_date,
                           actual_start_date, actual_end_date,
                           baseline_start_date, baseline_end_date,
                           notes, sort_order,
                           wbs_number, progress, budget, rag
                       ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (new_proj_id,
                     task_name,
                     task.get("type")   or "task",
                     task.get("status") or "not-started",
                     task.get("start_date"),
                     task.get("end_date"),
                     task.get("duration_days"),
                     1 if task.get("is_firm_date") else 0,
                     task.get("actual_start_date"),
                     task.get("actual_end_date"),
                     task.get("baseline_start_date"),
                     task.get("baseline_end_date"),
                     task.get("notes"),
                     task.get("sort_order") or 0,
                     task.get("wbs_number"),
                     float(task.get("progress") or 0.0),
                     task.get("budget"),
                     task.get("rag"))
                )
                new_task_id = cur.lastrowid
                if old_task_id is not None:
                    task_map[old_task_id] = new_task_id
                tasks_imported += 1

                # Assignees
                for old_pid in (task.get("assignee_ids") or []):
                    new_pid = people_map.get(old_pid)
                    if new_pid:
                        conn.execute(
                            "INSERT OR IGNORE INTO task_people (task_id, person_id) VALUES (?,?)",
                            (new_task_id, new_pid)
                        )

                # Checklist / note items
                for item in (task.get("items") or []):
                    conn.execute(
                        """INSERT INTO task_items
                           (task_id, content, item_type, is_complete, sort_order, value)
                           VALUES (?,?,?,?,?,?)""",
                        (new_task_id,
                         item.get("content") or "",
                         item.get("item_type") or "note",
                         1 if item.get("is_complete") else 0,
                         item.get("sort_order") or 0,
                         item.get("value"))
                    )

            # Tasks — second pass: remap parent_id using task_map
            for task in proj.get("tasks", []):
                old_parent_id = task.get("parent_id")
                old_task_id   = task.get("id")
                if old_parent_id is not None and old_task_id is not None:
                    new_task_id   = task_map.get(old_task_id)
                    new_parent_id = task_map.get(old_parent_id)
                    if new_task_id and new_parent_id:
                        conn.execute(
                            "UPDATE tasks SET parent_id = ? WHERE id = ?",
                            (new_parent_id, new_task_id)
                        )

            # Dependencies — insert after all tasks exist so FK constraints hold
            for dep in (proj.get("dependencies") or []):
                new_pred = task_map.get(dep.get("predecessor_id"))
                new_succ = task_map.get(dep.get("successor_id"))
                if new_pred and new_succ and new_pred != new_succ:
                    conn.execute(
                        """INSERT OR IGNORE INTO dependencies
                           (predecessor_id, successor_id, type, lag_days)
                           VALUES (?,?,?,?)""",
                        (new_pred, new_succ,
                         dep.get("type")     or "FS",
                         dep.get("lag_days") or 0)
                    )

            # Phases — re-create using the new project ID (no ID remapping needed)
            for phase in (proj.get("phases") or []):
                phase_name = (phase.get("name") or "").strip()
                if not phase_name:
                    continue
                conn.execute(
                    """INSERT INTO phases (project_id, name, start_date, end_date, colour, sort_order)
                       VALUES (?,?,?,?,?,?)""",
                    (new_proj_id,
                     phase_name,
                     phase.get("start_date"),
                     phase.get("end_date"),
                     phase.get("colour") or "#6366f1",
                     phase.get("sort_order") or 0)
                )

        conn.commit()
        return jsonify({
            "ok": True,
            "projects_imported": projects_imported,
            "tasks_imported":    tasks_imported,
        })

    except Exception as exc:
        conn.rollback()
        logger.error("Import failed: %s", exc)
        abort(500, f"Import failed: {exc}")
